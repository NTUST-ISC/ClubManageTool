from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

import openpyxl

from os import getenv
from dotenv import load_dotenv
from time import sleep

load_dotenv()

class Member():
    def __init__(self, _student_id: str, _name: str, _job: str = None) -> None:
        self.student_id = _student_id
        self.name = _name
        self.job = _job or "社員"
    
    def __str__(self) -> str:
        return f"({self.student_id}, {self.name}, {self.job})"
    
    def __repr__(self) -> str:
        return f"({self.student_id}, {self.name}, {self.job})"

def WaitForLoad(time: float = 0.2):
    def decorator(func):
        def wrapper(*args):
            func(*args)
            sleep(time)
        return wrapper
    return decorator

class WebControler():
    def __init__(self, driver_option):
        self.main_url = "https://clubs.ntust.edu.tw/Login/"
        self.member_manage_url = "https://clubs.ntust.edu.tw/ClubManagement/MemberList"
        self.driver = Chrome(options=driver_option)

    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_value, traceback):
        self.driver.close()

    @WaitForLoad()
    def click_button(self, selector: str) -> None:
        self.driver.find_element(By.CSS_SELECTOR, selector).click()

    @WaitForLoad()
    def send_string(self, name: str, msg: str) -> None:
        self.driver.find_element(By.NAME, name).send_keys(msg)

    @WaitForLoad()
    def set_select_menu(self, name: str, index: int) -> None:
        Select(self.driver.find_element(By.NAME, name)).select_by_index(index)

    @WaitForLoad(0.5)
    def login(self, username: str, password: str) -> None:
        self.click_button("#accordion > div:nth-child(4) > div > h4 > a")
        self.send_string("username", username)
        self.send_string("password", password)
        self.click_button("#login-form > div.modal-footer > div > button")

    def add_new_member(self, member: Member) -> None:
        self.send_string("StudentID", member.student_id)
        self.send_string("Name", member.name)
        if member.job != "社員":
            self.set_select_menu("Identity", 0)
        self.send_string("Title", member.job)
        self.click_button("body > div.container > div:nth-child(2) > div.col-sm-9.col-md-9 > form:nth-child(1) > p:nth-child(8) > button")

    def login_to_member_manage(self) -> None:
        self.driver.get(self.main_url)
        self.login(getenv("USERNAME"), getenv("PASSWORD"))
        # change to member manage page
        self.driver.get(self.member_manage_url)

def add_new_member_TEST():
    driver_option = ChromeOptions()
    driver_option.add_argument("start-maximized")
    with WebControler(driver_option) as web_controler:
        web_controler.login_to_member_manage()
        # 普通社員
        test_member = Member("B10901001", "測試員")
        # 幹部
        test_member2 = Member("B10901002", "測試員2", "測試職位2")
        web_controler.add_new_member(test_member)
        web_controler.add_new_member(test_member2)

def login_to_member_manage_TEST():
    driver_option = ChromeOptions()
    driver_option.add_argument("start-maximized")
    with WebControler(driver_option) as web_controler:
        web_controler.login_to_member_manage()

def load_xlsx(file_path: str) -> list[Member]:
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(min_row=2, min_col=1)
    members = [Member(*[c.value for c in row]) for row in rows]
    return members

def load_xlsxTEST():
    filename = "NTUSTISC-member-112-2.xlsx"
    members = load_xlsx(filename)
    for member in members:
        print(member)

def main():
    driver_option = ChromeOptions()
    driver_option.add_argument("start-maximized")
    members = load_xlsx("NTUSTISC-member-112-2.xlsx")
    failed_members = []
    with WebControler(driver_option) as web_controler:
        web_controler.login_to_member_manage()
        for member in members:
            try:
                web_controler.add_new_member(member)
            except Exception:
                failed_members.append(member)

    if len(failed_members) == 0:
        print("All members are added successfully.")
    else:
        print(f'Failed: {len(failed_members)} members.')
        for member in failed_members:
            print(member)

if "__main__" == __name__:
    main()